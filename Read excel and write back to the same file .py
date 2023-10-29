#!/usr/bin/env python
# coding: utf-8

# In[23]:


from openpyxl import load_workbook

file_path = "C:\\Users\\Home\\Documents\\learning_python_excel.xlsx"
workbook = load_workbook(file_path)
worksheet1, worksheet2, worksheet3 = workbook['Sheet1'], workbook['Sheet2'], workbook['Sheet3']
values = [cell.value for cell in worksheet2[1]]

n = 1
for row in worksheet3.iter_rows(min_row=n, max_row=100):
    for cell in row:
        cell.value = (
            f"set device-group {values[0]} prerule base "
            f"{worksheet1.cell(row=cell.row + 1, column=1).value} "
            f"{worksheet1.cell(row=cell.row + 1, column=6).value}"
            f" plus {values[2]}"
        )
        n += 1


workbook.save(file_path)


# In[ ]:





# In[ ]:




